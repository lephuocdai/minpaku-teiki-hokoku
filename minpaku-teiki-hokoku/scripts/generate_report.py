#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_report.py — deterministic engine for the minpaku 定期報告
(住宅宿泊事業法 第14条 / 施行規則 第12条 bi-monthly report).

USAGE
    python3 scripts/generate_report.py <workdir> <period>

    <workdir>  folder that contains minpaku-config.yaml (and one subfolder
               per reporting period)
    <period>   two consecutive months, formatted YYYY-MM_YYYY-MM
               (example: 2026-06_2026-07)

READS
    <workdir>/minpaku-config.yaml                       (required)
    <workdir>/<period>/normalized_reservations.csv     (required; header-only
                                                         file = zero-stay period)
    <workdir>/nationality-overrides.csv                 (optional)
    <workdir>/<period>/kujo-log.csv                     (optional; Kyoto only)
    <workdir>/<other same-fiscal-year period folders>/normalized_reservations.csv
               (optional; used ONLY for the 180-day running total)

WRITES into <workdir>/<period>/output/
    report_data.json                    single source of truth (re-renderable)
    <M#########>_report.xlsx            one review workbook per 届出番号
    teiki-hokoku-upload_<period>.csv    portal bulk-upload CSV (cp932 + CRLF)
    click-guide_<period>.md             per-property web-form click guide
    validation_<period>.md              validation report (always written)

COUNTING RULES (officially confirmed; sources in references/legal-requirements.md)
    * 宿泊日数   = count of unique occupied dates per 届出番号, i.e. the UNION
                   across all listings mapped to that 届出番号. A night is
                   checkin .. checkout-1, clipped to the reporting period.
                   (The legal day is noon-to-noon; this engine approximates
                   it by calendar date, which is exact for normal stays.)
    * 宿泊者数   = total individuals = adults + children + INFANTS. No official
                   text excludes infants; this engine has NO exclusion toggle.
                   A stay straddling two periods counts its guests in EACH
                   period (official rule).
    * 延べ人数   = sum of (guests × clipped nights) per reservation.
    * cancelled reservations (status == cancelled) are excluded from counts.
    * status == unknown rows ARE counted but surfaced in the validation
      report — never silently dropped.
    * nationality residence rules (foreigner with a Japan address → 日本;
      Japanese national without a Japan address → その他) are applied
      UPSTREAM by the skill dialogue; this engine receives final per-guest
      category values and only re-normalizes spellings as a safety net.

DEPENDENCIES
    Standard library + openpyxl (only needed when review_xlsx output is on).
"""

import csv
import calendar as cal_mod
import json
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta

# nationality_map.py lives in the same scripts/ folder. When this file is run
# as "python3 scripts/generate_report.py" sys.path[0] is already scripts/,
# but insert explicitly so importing this module from elsewhere also works.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import nationality_map as natmap

NATIONALITY_CATEGORIES = natmap.NATIONALITY_CATEGORIES
CATEGORY_SET = natmap.CATEGORY_SET

# ───────────────────────────── constants ─────────────────────────────

# Exact required header of the normalized reservations CSV (the seam between
# the skill dialogue, which maps ANY raw OTA export into this shape, and this
# deterministic engine, which does ALL counting).
NORMALIZED_HEADER = [
    "property_key", "listing_name", "source", "confirmation_code",
    "guest_name", "checkin", "checkout", "adults", "children", "infants",
    "status", "nationalities",
]

OVERRIDES_HEADER = ["confirmation_code", "mode", "nationalities"]
KUJO_HEADER = ["date", "time", "content", "response"]

VALID_STATUSES = {"confirmed", "cancelled", "unknown"}

# Language of generated documents; set from config in main(). ja is default.
LANG = "ja"


def T(ja, en):
    """Tiny i18n helper: pick the Japanese or English variant per config."""
    return ja if LANG == "ja" else en


def fatal(msg):
    """Unrecoverable input problem: print loudly and exit non-zero."""
    sys.stderr.write("ERROR: " + msg + "\n")
    sys.exit(2)


# ───────────────────────── period handling ───────────────────────────

PERIOD_RE = re.compile(r"^(\d{4})-(\d{2})_(\d{4})-(\d{2})$")
DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def parse_iso_date(s):
    """Strict YYYY-MM-DD parse; returns date or None."""
    s = str(s).strip()
    if not DATE_RE.match(s):
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def month_last_day(y, m):
    return date(y, m, cal_mod.monthrange(y, m)[1])


def parse_period(period_key):
    """'2026-06_2026-07' -> dict with start/end dates and portal labels.

    The two months must be consecutive. Standard reporting periods start on
    an even month (Dec-Jan, Feb-Mar, Apr-May, Jun-Jul, Aug-Sep, Oct-Nov);
    a non-standard period is allowed but flagged later.
    """
    m = PERIOD_RE.match(period_key)
    if not m:
        fatal("period must be YYYY-MM_YYYY-MM (e.g. 2026-06_2026-07), got %r"
              % period_key)
    sy, sm, ey, em = map(int, m.groups())
    if not (1 <= sm <= 12 and 1 <= em <= 12):
        fatal("period %r contains an invalid month" % period_key)
    if sy * 12 + sm + 1 != ey * 12 + em:
        fatal("period %r must cover two CONSECUTIVE months" % period_key)

    start = date(sy, sm, 1)
    end = month_last_day(ey, em)
    # Japanese fiscal year starts in April: Feb-Mar 2026 belongs to 2025年度.
    fiscal_year = sy if sm >= 4 else sy - 1
    return {
        "key": period_key,
        "start": start,
        "end": end,
        "fiscal_year": fiscal_year,
        "fiscal_year_start": date(fiscal_year, 4, 1),
        # 操作手順書 v1.7 serialization of 報告期間, e.g. 2026年度06月～07月.
        # NOTE the tilde is U+FF5E (FULLWIDTH TILDE) — the only tilde that
        # round-trips through cp932; U+301C WAVE DASH does not encode.
        "label_portal": "%d年度%02d月～%02d月" % (fiscal_year, sm, em),
        # 電子宿泊者名簿 v2.0 serialization, e.g. 2018_4-5 (non-padded months).
        # UNVERIFIED for periods whose calendar year differs from the fiscal
        # year (Feb-Mar): the manual only shows an April example. We use the
        # calendar year of the first month.
        "label_meibo": "%d_%d-%d" % (sy, sm, em),
        "is_standard": sm in (2, 4, 6, 8, 10, 12),
    }


# ─────────────────────── 届出番号 normalization ───────────────────────

# Accept 第M130000000号 / M130000000 / 第M130000000 / M130000000号, and
# tolerate full-width digits / full-width M (common when typed in a JA IME).
_FULLWIDTH = str.maketrans("０１２３４５６７８９Ｍｍ", "0123456789Mm")
TODOKEDE_RE = re.compile(r"^(?:第)?([Mm]\d{9})(?:号)?$")


def normalize_todokede(raw):
    """Return the bare canonical form 'M#########', or None if unparseable."""
    s = str(raw).strip().translate(_FULLWIDTH).replace(" ", "").replace("　", "")
    m = TODOKEDE_RE.match(s)
    if not m:
        return None
    return m.group(1).upper()


def display_todokede(bare):
    """Bare 'M#########' -> official display form '第M#########号'."""
    return "第%s号" % bare


# ─────────────────── minimal YAML subset parser ───────────────────────
# We deliberately avoid a third-party YAML library (openpyxl is the only
# allowed dependency). minpaku-config.yaml uses a small, documented subset:
# scalars, nested mappings, lists of mappings ("- key: value" style), and
# inline flow mappings ("{ ota: airbnb, account: a, listing_name: x }").
# Anything outside the subset raises a clear error instead of misparsing.


def _strip_comment(line):
    """Remove a trailing '# comment' that is outside quotes.

    Per YAML, '#' starts a comment at line start or when preceded by
    whitespace; '#' inside quoted strings is preserved.
    """
    in_single = in_double = False
    for i, ch in enumerate(line):
        if ch == "'" and not in_double:
            in_single = not in_single
        elif ch == '"' and not in_single:
            in_double = not in_double
        elif ch == "#" and not in_single and not in_double:
            if i == 0 or line[i - 1] in " \t":
                return line[:i].rstrip()
    return line.rstrip()


def _parse_scalar(tok):
    tok = tok.strip()
    if len(tok) >= 2 and tok[0] == '"' and tok[-1] == '"':
        return tok[1:-1]
    if len(tok) >= 2 and tok[0] == "'" and tok[-1] == "'":
        return tok[1:-1]
    low = tok.lower()
    if low in ("true", "yes", "on"):
        return True
    if low in ("false", "no", "off"):
        return False
    if low in ("null", "~", ""):
        return None
    if tok.isdigit():
        return int(tok)
    return tok


def _split_top_level(s, sep=","):
    """Split on `sep` outside quotes and outside {}/[] nesting."""
    parts, buf = [], []
    in_single = in_double = False
    depth = 0
    for ch in s:
        if ch == "'" and not in_double:
            in_single = not in_single
        elif ch == '"' and not in_single:
            in_double = not in_double
        if ch == sep and not in_single and not in_double and depth == 0:
            parts.append("".join(buf))
            buf = []
            continue
        if not in_single and not in_double:
            if ch in "{[":
                depth += 1
            elif ch in "}]":
                depth -= 1
        buf.append(ch)
    parts.append("".join(buf))
    return parts


def _parse_flow_map(s):
    """'{ ota: airbnb, account: "a", listing_name: "x" }' -> dict."""
    s = s.strip()
    if not (s.startswith("{") and s.endswith("}")):
        raise ValueError("expected inline mapping {...}, got: %r" % s)
    inner = s[1:-1].strip()
    result = {}
    if not inner:
        return result
    for part in _split_top_level(inner):
        key, sep, val = part.partition(":")
        if not sep:
            raise ValueError("bad inline mapping entry: %r" % part)
        result[key.strip().strip("\"'")] = _parse_scalar(val)
    return result


def _parse_flow_list(s):
    s = s.strip()
    if not (s.startswith("[") and s.endswith("]")):
        raise ValueError("expected inline list [...], got: %r" % s)
    inner = s[1:-1].strip()
    if not inner:
        return []
    return [_parse_scalar(p) for p in _split_top_level(inner)]


def _parse_mapping(lines, idx, indent):
    result = {}
    while idx < len(lines):
        ind, content, lineno = lines[idx]
        if ind < indent:
            break
        if ind > indent:
            raise ValueError("line %d: unexpected indentation" % lineno)
        if content == "-" or content.startswith("- "):
            break  # a sequence at this level belongs to the parent key
        key, sep, rest = content.partition(":")
        if not sep:
            raise ValueError("line %d: expected 'key: value', got %r"
                             % (lineno, content))
        key = key.strip().strip("\"'")
        rest = rest.strip()
        if rest == "":
            # Nested block (mapping or sequence) on the following lines.
            if idx + 1 < len(lines) and lines[idx + 1][0] > indent:
                child_indent = lines[idx + 1][0]
                value, idx = _parse_block(lines, idx + 1, child_indent)
            else:
                value = None
                idx += 1
            result[key] = value
        elif rest.startswith("{"):
            result[key] = _parse_flow_map(rest)
            idx += 1
        elif rest.startswith("["):
            result[key] = _parse_flow_list(rest)
            idx += 1
        else:
            result[key] = _parse_scalar(rest)
            idx += 1
    return result, idx


def _parse_sequence(lines, idx, indent):
    result = []
    while idx < len(lines):
        ind, content, lineno = lines[idx]
        if ind != indent or not (content == "-" or content.startswith("- ")):
            break
        after = content[1:]  # text after the dash
        rest = after.strip()
        if rest == "":
            # "-" alone: the item is a nested block on the following lines.
            if idx + 1 < len(lines) and lines[idx + 1][0] > indent:
                child_indent = lines[idx + 1][0]
                value, idx = _parse_block(lines, idx + 1, child_indent)
                result.append(value)
            else:
                result.append(None)
                idx += 1
        elif rest.startswith("{"):
            result.append(_parse_flow_map(rest))
            idx += 1
        elif ":" in rest and not rest.startswith(("\"", "'")):
            # "- key: value" — the item is a mapping whose first entry sits
            # on the dash line. Continuation keys must be indented to the
            # first key's column (dash column + 1 + spaces after the dash).
            item_indent = indent + 1 + (len(after) - len(after.lstrip()))
            lines[idx] = (item_indent, rest, lineno)
            value, idx = _parse_mapping(lines, idx, item_indent)
            result.append(value)
        else:
            result.append(_parse_scalar(rest))
            idx += 1
    return result, idx


def _parse_block(lines, idx, indent):
    if lines[idx][1] == "-" or lines[idx][1].startswith("- "):
        return _parse_sequence(lines, idx, indent)
    return _parse_mapping(lines, idx, indent)


def parse_simple_yaml(text):
    """Parse the documented minpaku-config.yaml subset into dicts/lists."""
    lines = []
    for lineno, raw in enumerate(text.splitlines(), 1):
        stripped_line = _strip_comment(raw)
        if not stripped_line.strip():
            continue
        indent = len(stripped_line) - len(stripped_line.lstrip(" "))
        if "\t" in stripped_line[:indent + 1]:
            raise ValueError("line %d: tabs are not allowed in indentation"
                             % lineno)
        lines.append((indent, stripped_line.strip(), lineno))
    if not lines:
        return {}
    value, next_idx = _parse_block(lines, 0, lines[0][0])
    if next_idx != len(lines):
        raise ValueError("line %d: could not parse rest of file (check "
                         "indentation)" % lines[next_idx][2])
    return value


# ───────────────────────── config loading ─────────────────────────────


def load_config(path):
    """Load and validate minpaku-config.yaml; returns a normalized dict."""
    if not os.path.isfile(path):
        fatal("config not found: %s\n(create minpaku-config.yaml in the "
              "work folder — see the skill's onboarding)" % path)
    with open(path, "r", encoding="utf-8-sig") as f:
        try:
            raw = parse_simple_yaml(f.read())
        except ValueError as e:
            fatal("could not parse %s: %s" % (path, e))
    if not isinstance(raw, dict):
        fatal("%s: top level must be a mapping" % path)

    cfg = {}
    cfg["language"] = str(raw.get("language") or "ja").strip().lower()
    if cfg["language"] not in ("ja", "en"):
        fatal("config language must be 'ja' or 'en', got %r" % cfg["language"])
    cfg["operator_name"] = str(raw.get("operator_name") or "").strip()
    cfg["csv_variant"] = str(raw.get("csv_variant") or "sousa-tejunsho").strip()
    if cfg["csv_variant"] not in ("sousa-tejunsho", "denshi-meibo"):
        fatal("config csv_variant must be 'sousa-tejunsho' or 'denshi-meibo',"
              " got %r" % cfg["csv_variant"])

    out = raw.get("output") or {}
    if not isinstance(out, dict):
        fatal("config 'output' must be a mapping")
    cfg["output"] = {
        "portal_csv": bool(out.get("portal_csv", True)),
        "click_guide": bool(out.get("click_guide", True)),
        "review_xlsx": bool(out.get("review_xlsx", True)),
    }

    props_raw = raw.get("properties")
    if not isinstance(props_raw, list) or not props_raw:
        fatal("config 'properties' must be a non-empty list")
    properties = {}
    for i, p in enumerate(props_raw, 1):
        if not isinstance(p, dict):
            fatal("config properties[%d] must be a mapping" % i)
        bare = normalize_todokede(p.get("todokede_no") or "")
        if not bare:
            fatal("config properties[%d]: todokede_no %r is not a valid "
                  "届出番号 (expected 第M+9桁+号, with or without 第/号)"
                  % (i, p.get("todokede_no")))
        if bare in properties:
            fatal("config: duplicate 届出番号 %s" % display_todokede(bare))
        juri = None
        if p.get("juri_date"):
            juri = parse_iso_date(p["juri_date"])
            if not juri:
                fatal("config properties[%d]: juri_date %r must be YYYY-MM-DD"
                      % (i, p["juri_date"]))
        listings = []
        for li, l in enumerate(p.get("listings") or [], 1):
            if not isinstance(l, dict):
                fatal("config properties[%d].listings[%d] must be a mapping"
                      % (i, li))
            listings.append({
                "ota": str(l.get("ota") or "").strip(),
                "account": str(l.get("account") or "").strip(),
                "listing_name": str(l.get("listing_name") or "").strip(),
            })
        properties[bare] = {
            "todokede": bare,
            "todokede_display": display_todokede(bare),
            "name": str(p.get("name") or bare).strip(),
            "municipality": str(p.get("municipality") or "").strip(),
            "kyoto": bool(p.get("kyoto", False)),
            "juri_date": juri,
            "listings": listings,
        }
    cfg["properties"] = properties
    return cfg


# ─────────────────────── findings (validation) ────────────────────────


class Findings:
    """Collector for validation results at three severities.

    ERROR   = do not submit until fixed.
    WARNING = review before submitting.
    INFO    = for the record.
    """

    def __init__(self):
        self.errors = []
        self.warnings = []
        self.infos = []

    def error(self, msg):
        self.errors.append(msg)

    def warning(self, msg):
        self.warnings.append(msg)

    def info(self, msg):
        self.infos.append(msg)


# ───────────────────── reservations CSV loading ───────────────────────


def check_header(actual, expected, path):
    actual = [c.strip().lstrip("﻿") for c in actual]
    if actual != expected:
        fatal("%s: header mismatch.\n  expected: %s\n  got:      %s"
              % (path, ",".join(expected), ",".join(actual)))


def load_reservations(path, period, known_props, findings):
    """Read normalized_reservations.csv → list of in-period reservation dicts.

    Also returns counters used by the validation report. Bad rows are
    surfaced as ERRORs and skipped; out-of-period rows are counted as INFO;
    cancelled rows are excluded from counts (that is the documented rule,
    not silent dropping).
    """
    if not os.path.isfile(path):
        fatal("reservations file not found: %s\n"
              "For a genuine zero-stay period, create the file with the "
              "header row only — that makes 'no stays' an explicit statement "
              "instead of a missing input." % path)

    records = []
    stats = {"rows": 0, "cancelled": 0, "out_of_period": 0, "bad": 0,
             "unknown_status": 0, "unknown_property": 0}

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        try:
            header = next(reader)
        except StopIteration:
            fatal("%s is empty (expected at least the header row)" % path)
        check_header(header, NORMALIZED_HEADER, path)

        for lineno, row in enumerate(reader, 2):
            if not any(cell.strip() for cell in row):
                continue  # fully blank line
            stats["rows"] += 1
            if len(row) != len(NORMALIZED_HEADER):
                stats["bad"] += 1
                findings.error(T(
                    "%s 行%d: 列数が%dではありません (%d列) — 行をスキップ"
                    % (os.path.basename(path), lineno,
                       len(NORMALIZED_HEADER), len(row)),
                    "%s line %d: expected %d columns, got %d — row skipped"
                    % (os.path.basename(path), lineno,
                       len(NORMALIZED_HEADER), len(row))))
                continue
            r = dict(zip(NORMALIZED_HEADER, [c.strip() for c in row]))

            # ── 届出番号 (property_key) ──
            bare = normalize_todokede(r["property_key"])
            if not bare:
                stats["bad"] += 1
                findings.error(T(
                    "行%d: property_key %r が届出番号として解釈できません — 行をスキップ"
                    % (lineno, r["property_key"]),
                    "line %d: property_key %r is not a valid 届出番号 — row "
                    "skipped" % (lineno, r["property_key"])))
                continue
            if bare not in known_props:
                stats["unknown_property"] += 1
                findings.error(T(
                    "行%d: 届出番号 %s は minpaku-config.yaml に存在しません "
                    "(予約 %s) — 集計から除外。設定に物件を追加してください"
                    % (lineno, display_todokede(bare), r["confirmation_code"]),
                    "line %d: 届出番号 %s not found in minpaku-config.yaml "
                    "(reservation %s) — excluded from totals. Add the "
                    "property to the config."
                    % (lineno, display_todokede(bare), r["confirmation_code"])))
                continue

            # ── dates ──
            checkin = parse_iso_date(r["checkin"])
            checkout = parse_iso_date(r["checkout"])
            if not checkin or not checkout:
                stats["bad"] += 1
                findings.error(T(
                    "行%d (%s): checkin/checkout はYYYY-MM-DD形式が必要 — 行をスキップ"
                    % (lineno, r["confirmation_code"]),
                    "line %d (%s): checkin/checkout must be YYYY-MM-DD — row "
                    "skipped" % (lineno, r["confirmation_code"])))
                continue
            if checkout <= checkin:
                stats["bad"] += 1
                findings.error(T(
                    "行%d (%s): checkout がcheckin以前です (%s → %s) — 行をスキップ"
                    % (lineno, r["confirmation_code"], checkin, checkout),
                    "line %d (%s): checkout is not after checkin (%s → %s) — "
                    "row skipped"
                    % (lineno, r["confirmation_code"], checkin, checkout)))
                continue

            # ── party size ──
            counts = {}
            bad_int = False
            for fld in ("adults", "children", "infants"):
                v = r[fld]
                if v == "":
                    counts[fld] = 0
                elif v.lstrip("-").isdigit() and int(v) >= 0:
                    counts[fld] = int(v)
                else:
                    bad_int = True
            if bad_int:
                stats["bad"] += 1
                findings.error(T(
                    "行%d (%s): adults/children/infants は0以上の整数が必要 — 行をスキップ"
                    % (lineno, r["confirmation_code"]),
                    "line %d (%s): adults/children/infants must be "
                    "non-negative integers — row skipped"
                    % (lineno, r["confirmation_code"])))
                continue
            # 宿泊者数 counts EVERY individual: adults + children + infants.
            guests = counts["adults"] + counts["children"] + counts["infants"]
            if guests > 500:
                stats["bad"] += 1
                findings.error(T(
                    "行%d (%s): 宿泊者数が%d人です — 500人を超えており異常値です。"
                    "normalized_reservations.csv を修正してください — 行をスキップ"
                    % (lineno, r["confirmation_code"], guests),
                    "line %d (%s): party size is %d — exceeds 500 and is "
                    "implausible; fix normalized_reservations.csv — row "
                    "skipped" % (lineno, r["confirmation_code"], guests)))
                continue
            if guests == 0:
                findings.warning(T(
                    "行%d (%s): 宿泊者数が0人です — データを確認してください"
                    % (lineno, r["confirmation_code"]),
                    "line %d (%s): party size is 0 — check the source data"
                    % (lineno, r["confirmation_code"])))

            # ── status ──
            status = r["status"].lower()
            if status not in VALID_STATUSES:
                findings.warning(T(
                    "行%d (%s): status %r は confirmed/cancelled/unknown の"
                    "いずれでもありません — unknown として扱い、計上します"
                    % (lineno, r["confirmation_code"], r["status"]),
                    "line %d (%s): status %r is not one of "
                    "confirmed/cancelled/unknown — treated as unknown and "
                    "COUNTED" % (lineno, r["confirmation_code"], r["status"])))
                status = "unknown"
            if status == "cancelled":
                stats["cancelled"] += 1
                continue  # excluded from all counts (documented rule)
            if status == "unknown":
                stats["unknown_status"] += 1
                # Surfaced (not dropped): the row IS counted, but the user
                # must confirm it was a real stay before filing.
                findings.warning(T(
                    "normalized_reservations.csv 行%d (%s): ステータス不明 — "
                    "実在した宿泊なら計上のままでOK、実在しない予約なら"
                    "この行の削除が必要です（スキルに「行%dを削除して」と"
                    "伝えてください。現在は計上しています）"
                    % (lineno, r["confirmation_code"], lineno),
                    "normalized_reservations.csv line %d (%s): status "
                    "unknown — currently COUNTED; keep if it was a real "
                    "stay, otherwise the row must be removed (just tell the "
                    "skill to delete line %d)"
                    % (lineno, r["confirmation_code"], lineno)))

            # ── clip the stay to the reporting period ──
            last_night = checkout - timedelta(days=1)
            clip_start = max(checkin, period["start"])
            clip_end = min(last_night, period["end"])
            if clip_start > clip_end:
                stats["out_of_period"] += 1
                continue  # reservation entirely outside this period
            clipped_nights = (clip_end - clip_start).days + 1

            # ── nationalities: semicolon-joined per-guest final values ──
            raw_nats = [s.strip() for s in r["nationalities"].split(";")] \
                if r["nationalities"] != "" else []
            if len(raw_nats) > guests:
                findings.warning(T(
                    "行%d (%s): 国籍の数(%d)が宿泊者数(%d)より多い — 超過分は無視"
                    % (lineno, r["confirmation_code"], len(raw_nats), guests),
                    "line %d (%s): %d nationalities for %d guests — extras "
                    "ignored" % (lineno, r["confirmation_code"],
                                 len(raw_nats), guests)))
                raw_nats = raw_nats[:guests]
            while len(raw_nats) < guests:
                raw_nats.append("")  # blank slot = unresolved

            nats, nat_sources = [], []
            for raw_val in raw_nats:
                if raw_val == "":
                    nats.append(None)
                    nat_sources.append("unresolved")
                elif raw_val in CATEGORY_SET:
                    nats.append(raw_val)
                    nat_sources.append("csv")
                else:
                    cat = natmap.normalize_nationality(raw_val)
                    if cat:
                        nats.append(cat)
                        nat_sources.append("csv(%s→%s)" % (raw_val, cat))
                    else:
                        nats.append(None)
                        nat_sources.append("unresolved(%s)" % raw_val)
                        findings.warning(T(
                            "行%d (%s): 国籍 %r を22分類に変換できません — 未確定として扱います"
                            % (lineno, r["confirmation_code"], raw_val),
                            "line %d (%s): nationality %r not recognized — "
                            "treated as unresolved"
                            % (lineno, r["confirmation_code"], raw_val)))

            records.append({
                "todokede": bare,
                "listing_name": r["listing_name"],
                "source": r["source"],
                "code": r["confirmation_code"],
                "guest_name": r["guest_name"],
                "checkin": checkin,
                "checkout": checkout,
                "clip_start": clip_start,
                "clip_end": clip_end,
                "clipped_nights": clipped_nights,
                "adults": counts["adults"],
                "children": counts["children"],
                "infants": counts["infants"],
                "guests": guests,
                "person_nights": guests * clipped_nights,
                "status": status,
                "nats": nats,
                "nat_sources": nat_sources,
            })
    return records, stats


# ───────────────────── nationality overrides ──────────────────────────


def load_overrides(path, findings):
    """Read nationality-overrides.csv → {code: (mode, [values])}."""
    overrides = {}
    if not os.path.isfile(path):
        return overrides
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        try:
            header = next(reader)
        except StopIteration:
            return overrides
        check_header(header, OVERRIDES_HEADER, path)
        for lineno, row in enumerate(reader, 2):
            if not any(c.strip() for c in row):
                continue
            if len(row) != 3:
                findings.warning(T(
                    "overrides 行%d: 列数が3ではありません — スキップ" % lineno,
                    "overrides line %d: expected 3 columns — skipped" % lineno))
                continue
            code, mode, values = (c.strip() for c in row)
            mode = mode.lower()
            if mode not in ("fill", "force"):
                findings.warning(T(
                    "overrides 行%d (%s): mode は fill|force のみ (%r) — スキップ"
                    % (lineno, code, mode),
                    "overrides line %d (%s): mode must be fill|force, got %r "
                    "— skipped" % (lineno, code, mode)))
                continue
            vals = [v.strip() for v in values.split(";") if v.strip()]
            if not vals:
                findings.warning(T(
                    "overrides 行%d (%s): nationalities が空 — スキップ"
                    % (lineno, code),
                    "overrides line %d (%s): empty nationalities — skipped"
                    % (lineno, code)))
                continue
            if code in overrides:
                findings.warning(T(
                    "overrides: %s が複数回定義 — 後の行を使用" % code,
                    "overrides: %s defined more than once — last row wins"
                    % code))
            overrides[code] = (mode, vals)
    return overrides


def apply_overrides(records, overrides, findings):
    """Apply fill/force overrides to loaded reservations, in place.

    fill : fills BLANK slots only. One value = every blank slot gets it;
           multiple values are consumed per blank slot in order (last value
           repeats if there are more blanks than values).
    force: replaces the ENTIRE per-guest list. One value = all guests;
           multiple values are aligned per guest (mismatch → warning, last
           value repeats / extras ignored).
    Override values are themselves normalized; unrecognized values are
    refused (the slot stays unresolved) — an override must never inject a
    value the portal form cannot accept.
    """
    used = set()
    for rec in records:
        if rec["code"] not in overrides:
            continue
        used.add(rec["code"])
        mode, raw_vals = overrides[rec["code"]]

        cats = []
        for v in raw_vals:
            cat = v if v in CATEGORY_SET else natmap.normalize_nationality(v)
            if not cat:
                findings.warning(T(
                    "override (%s): 値 %r を22分類に変換できません — この値は無視"
                    % (rec["code"], v),
                    "override (%s): value %r not recognized — ignored"
                    % (rec["code"], v)))
                continue
            cats.append(cat)
        if not cats:
            continue

        if mode == "force":
            if len(cats) not in (1, rec["guests"]):
                findings.warning(T(
                    "override force (%s): 値%d件と宿泊者%d人が不一致 — 末尾値で補完/切り詰め"
                    % (rec["code"], len(cats), rec["guests"]),
                    "override force (%s): %d values vs %d guests — padded "
                    "with last value / truncated"
                    % (rec["code"], len(cats), rec["guests"])))
            new = [cats[i] if i < len(cats) else cats[-1]
                   for i in range(rec["guests"])]
            rec["nats"] = new
            rec["nat_sources"] = ["override:force"] * rec["guests"]
        else:  # fill — blanks only, existing values untouched
            vi = 0
            for i in range(rec["guests"]):
                if rec["nats"][i] is None:
                    val = cats[vi] if vi < len(cats) else cats[-1]
                    if len(cats) > 1:
                        vi += 1
                    rec["nats"][i] = val
                    rec["nat_sources"][i] = "override:fill"

    for code in overrides:
        if code not in used:
            findings.info(T(
                "override %s: この期間の予約に該当なし (期間外またはコード相違)" % code,
                "override %s: no matching reservation in this period (out of "
                "period, or code mismatch)" % code))


# ─────────────────────── duplicate detection ──────────────────────────


def detect_duplicates(records, findings):
    """Duplicate confirmation codes + cross-source duplicate stays.

    Cross-source rule (from the design contract): same 届出番号 + same
    checkin/checkout arriving from DIFFERENT sources, with the same guest
    name OR the same party size → probably the same booking imported twice.
    We WARN and keep both rows (never silently drop) so the user decides.
    """
    by_code = defaultdict(list)
    for rec in records:
        if rec["code"]:
            by_code[rec["code"]].append(rec)
    for code, recs in by_code.items():
        if len(recs) > 1:
            findings.warning(T(
                "確認コード %s が%d回出現 — 同一予約なら1行に統合してください"
                "(現在は重複計上されています)" % (code, len(recs)),
                "confirmation code %s appears %d times — if it is the same "
                "booking, merge to one row (currently double-counted)"
                % (code, len(recs))))

    by_stay = defaultdict(list)
    for rec in records:
        by_stay[(rec["todokede"], rec["checkin"], rec["checkout"])].append(rec)
    for (bare, ci, co), recs in by_stay.items():
        if len(recs) < 2:
            continue
        for i in range(len(recs)):
            for j in range(i + 1, len(recs)):
                a, b = recs[i], recs[j]
                if a["source"] == b["source"]:
                    continue
                same_name = (a["guest_name"].strip().lower()
                             == b["guest_name"].strip().lower()
                             and a["guest_name"].strip() != "")
                same_party = a["guests"] == b["guests"]
                if same_name or same_party:
                    findings.warning(T(
                        "%s: %s〜%s の滞在が複数ソースに存在 (%s:%s と %s:%s) — "
                        "同一予約の二重取り込みの可能性。どちらが正しいか確認し、"
                        "normalized_reservations.csv から片方の行の削除が必要です"
                        "（スキルに「%s と %s のどちらを残すか」を伝えれば"
                        "削除してもらえます）"
                        % (display_todokede(bare), ci, co,
                           a["source"], a["code"], b["source"], b["code"],
                           a["code"], b["code"]),
                        "%s: stay %s–%s present from multiple sources "
                        "(%s:%s and %s:%s) — possible double import of the "
                        "same booking; verify which is real, then one row "
                        "must be removed from normalized_reservations.csv "
                        "(just tell the skill which of %s / %s to keep)"
                        % (display_todokede(bare), ci, co,
                           a["source"], a["code"], b["source"], b["code"],
                           a["code"], b["code"])))


# ─────────────────── 180-day fiscal running total ─────────────────────


def scan_fiscal_occupancy(workdir, period, known_props, current_records):
    """Occupied-date UNION per property across the fiscal year so far.

    The 180-day cap year runs noon Apr 1 → noon Apr 1 (規則第3条); we
    approximate by calendar date. Data sources: the CURRENT period's records
    plus any sibling <workdir>/<YYYY-MM_YYYY-MM>/normalized_reservations.csv
    folders whose period starts inside the same fiscal year and before this
    one. The result is only as complete as the folders present — the caller
    must say so (fail-loud) in the validation report.
    """
    fy_start = period["fiscal_year_start"]
    window_end = period["end"]
    dates_by_prop = defaultdict(set)
    sources = [period["key"] + " (current)"]

    def add_stay(bare, checkin, checkout):
        last_night = checkout - timedelta(days=1)
        d = max(checkin, fy_start)
        end = min(last_night, window_end)
        while d <= end:
            dates_by_prop[bare].add(d)
            d += timedelta(days=1)

    for rec in current_records:
        add_stay(rec["todokede"], rec["checkin"], rec["checkout"])

    # Sibling period folders in the same fiscal year, before this period.
    try:
        entries = sorted(os.listdir(workdir))
    except OSError:
        entries = []
    for name in entries:
        m = PERIOD_RE.match(name)
        if not m or name == period["key"]:
            continue
        sy, sm = int(m.group(1)), int(m.group(2))
        sib_start = date(sy, sm, 1)
        if not (fy_start <= sib_start < period["start"]):
            continue
        sib_csv = os.path.join(workdir, name, "normalized_reservations.csv")
        if not os.path.isfile(sib_csv):
            continue
        sources.append(name)
        with open(sib_csv, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row is None:
                    continue
                bare = normalize_todokede(row.get("property_key") or "")
                if not bare or bare not in known_props:
                    continue
                if (row.get("status") or "").strip().lower() == "cancelled":
                    continue
                ci = parse_iso_date(row.get("checkin") or "")
                co = parse_iso_date(row.get("checkout") or "")
                if ci and co and co > ci:
                    add_stay(bare, ci, co)

    return {bare: len(ds) for bare, ds in dates_by_prop.items()}, sources


# ─────────────────────── 苦情ログ (Kyoto) ─────────────────────────────


def load_kujo_log(path, period, findings):
    """Read kujo-log.csv (Kyoto complaint log) → list of in-period entries.

    京都市条例16条1項 adds 苦情の件数・日時・内容・対応状況 to the bi-monthly
    report. The log file has no per-property column (v1 limitation): it is
    attached to EVERY kyoto:true property, which is only correct when the
    operator has a single Kyoto property or keeps one log per work folder.
    """
    entries = []
    if not os.path.isfile(path):
        return entries
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        try:
            header = next(reader)
        except StopIteration:
            return entries
        check_header(header, KUJO_HEADER, path)
        for lineno, row in enumerate(reader, 2):
            if not any(c.strip() for c in row):
                continue
            if len(row) != 4:
                findings.warning(T(
                    "kujo-log 行%d: 列数が4ではありません — スキップ" % lineno,
                    "kujo-log line %d: expected 4 columns — skipped" % lineno))
                continue
            d = parse_iso_date(row[0])
            if not d:
                findings.warning(T(
                    "kujo-log 行%d: date はYYYY-MM-DD形式が必要 (%r) — スキップ"
                    % (lineno, row[0]),
                    "kujo-log line %d: date must be YYYY-MM-DD (%r) — skipped"
                    % (lineno, row[0])))
                continue
            if not (period["start"] <= d <= period["end"]):
                findings.warning(T(
                    "kujo-log 行%d: %s は報告期間外 — 今回の報告には含めません"
                    % (lineno, d),
                    "kujo-log line %d: %s is outside the reporting period — "
                    "not included this cycle" % (lineno, d)))
                continue
            entries.append({"date": d.isoformat(), "time": row[1].strip(),
                            "content": row[2].strip(),
                            "response": row[3].strip()})
    return entries


# ───────────────────── per-property aggregation ───────────────────────


def build_property_payload(prop, records, period, kujo_entries, findings):
    """Compute one 届出番号's report figures from its reservations.

    宿泊日数 is the UNION of occupied dates across every listing/row mapped
    to this 届出番号 — two listings occupied on the same date = 1 day.
    """
    occupied = set()
    total_guests = 0
    total_person_nights = 0
    nat_counts = defaultdict(int)
    unresolved = 0
    details = []
    review_items = []

    for rec in sorted(records, key=lambda r: (r["checkin"], r["code"])):
        d = rec["clip_start"]
        while d <= rec["clip_end"]:
            occupied.add(d)
            d += timedelta(days=1)
        # Straddling stays: guests count once in EACH period they touch
        # (official rule) — so no proration here, only night clipping.
        total_guests += rec["guests"]
        total_person_nights += rec["person_nights"]

        issues = []
        for nat in rec["nats"]:
            if nat is None:
                unresolved += 1
            else:
                nat_counts[nat] += 1
        n_unres = sum(1 for n in rec["nats"] if n is None)
        if n_unres:
            issues.append(T("国籍未確定 %d名" % n_unres,
                            "%d unresolved nationalities" % n_unres))
        if rec["status"] == "unknown":
            issues.append(T("ステータス不明(計上済み — 実在予約か確認)",
                            "status unknown (counted — verify it is a real "
                            "stay)"))

        details.append({
            "code": rec["code"],
            "guest_name": rec["guest_name"],
            "listing_name": rec["listing_name"],
            "source": rec["source"],
            "checkin": rec["checkin"].isoformat(),
            "checkout": rec["checkout"].isoformat(),
            "clip_start": rec["clip_start"].isoformat(),
            "clip_end": rec["clip_end"].isoformat(),
            "clipped_nights": rec["clipped_nights"],
            "adults": rec["adults"],
            "children": rec["children"],
            "infants": rec["infants"],
            "guests": rec["guests"],
            "person_nights": rec["person_nights"],
            "status": rec["status"],
            "nationalities": rec["nats"],
            "nat_sources": rec["nat_sources"],
            "issues": issues,
        })
        if issues:
            review_items.append({"code": rec["code"],
                                 "guest_name": rec["guest_name"],
                                 "guests": rec["guests"],
                                 "issues": issues})

    # Internal consistency asserts (these mirror the portal's own checks:
    # the web form derives 宿泊者数 from the nationality-grid sum).
    if sum(nat_counts.values()) + unresolved != total_guests:
        findings.error(T(
            "%s: 内部整合性エラー — 国籍合計+未確定(%d) ≠ 宿泊者数(%d)。バグの可能性、"
            "結果を使用しないでください"
            % (prop["todokede_display"],
               sum(nat_counts.values()) + unresolved, total_guests),
            "%s: internal consistency error — nationality sum + unresolved "
            "(%d) != guest total (%d). Possible bug; do not use this output"
            % (prop["todokede_display"],
               sum(nat_counts.values()) + unresolved, total_guests)))
    if unresolved:
        findings.error(T(
            "%s: 国籍未確定が%d名あります — 提出前に必ず解決してください "
            "(nationality-overrides.csv で補完可能)"
            % (prop["todokede_display"], unresolved),
            "%s: %d unresolved nationalities — resolve before filing "
            "(nationality-overrides.csv can fill them)"
            % (prop["todokede_display"], unresolved)))

    zero_report = len(details) == 0
    if zero_report:
        findings.info(T(
            "%s: この期間の宿泊実績なし — 0件報告を生成 (0件でも報告義務あり)"
            % prop["todokede_display"],
            "%s: no stays this period — zero report generated (a zero "
            "report is still legally required)" % prop["todokede_display"]))

    return {
        "todokede": prop["todokede"],
        "todokede_display": prop["todokede_display"],
        "name": prop["name"],
        "municipality": prop["municipality"],
        "kyoto": prop["kyoto"],
        "juri_date": prop["juri_date"].isoformat() if prop["juri_date"] else None,
        "zero_report": zero_report,
        "reservation_count": len(details),
        "shukuhaku_nissu": len(occupied),          # ① 宿泊日数
        "shukuhakusha_su": total_guests,           # ② 宿泊者数
        "nobe_ninzu": total_person_nights,         # ③ 延べ人数
        "nationality_counts": {c: nat_counts.get(c, 0)
                               for c in NATIONALITY_CATEGORIES},
        "unresolved_count": unresolved,
        "occupied_dates": sorted(d.isoformat() for d in occupied),
        "reservations": details,
        "review_items": review_items,
        "kujo": kujo_entries if prop["kyoto"] else None,
    }


# ───────────────────────── xlsx rendering ─────────────────────────────


def _safe_cell(v):
    # Neutralize spreadsheet formula injection from untrusted text
    # (Excel/Sheets execute cells starting with = + - @ and tab/CR variants).
    if isinstance(v, str) and v[:1] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + v
    return v


def render_property_xlsx(payload, period, output_dir):
    """Render one property's review workbook (canonical 事業実績報告 layout:
    calendar grid per month, 22-category 7-column table, summary row,
    明細 sheet, 確認ログ sheet, plus 苦情ログ for Kyoto)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        fatal("openpyxl is required for the review xlsx output.\n"
              "Install it with:  pip install openpyxl\n"
              "(or set output.review_xlsx: false in minpaku-config.yaml)")

    HEADER_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3",
                              fill_type="solid")
    OCCUPIED_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                fill_type="solid")
    ALERT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE",
                             fill_type="solid")
    CENTER = Alignment(horizontal="center", vertical="center")

    occupied = set(date.fromisoformat(d) for d in payload["occupied_dates"])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: 事業実績報告 (mirrors the government web form) ──
    ws1 = wb.create_sheet(title="事業実績報告")
    ws1["A1"] = "事業実績報告"
    ws1["A1"].font = Font(size=14, bold=True)
    ws1["A2"] = "届出番号"; ws1["B2"] = payload["todokede_display"]
    ws1["A3"] = "報告期間"; ws1["B3"] = period["label_portal"]
    ws1["A4"] = "物件名";   ws1["B4"] = _safe_cell(payload["name"])
    for coord in ("A2", "A3", "A4"):
        ws1[coord].font = Font(bold=True)

    ws1["A6"] = "宿泊日選択"
    ws1["A6"].font = Font(bold=True)

    # Weekly calendar grid, Sunday-first (matches the portal's calendar).
    weekday_jp = ["日", "月", "火", "水", "木", "金", "土"]
    cur_row = 7
    y, m = period["start"].year, period["start"].month
    for _ in range(2):  # a period is always exactly two months
        ws1.cell(row=cur_row, column=1,
                 value="%d年%d月" % (y, m)).font = Font(bold=True)
        cur_row += 1
        for i, wd in enumerate(weekday_jp):
            c = ws1.cell(row=cur_row, column=i + 1, value=wd)
            c.fill = HEADER_FILL
            c.alignment = CENTER
            c.font = Font(bold=True)
        cur_row += 1

        days_in_month = cal_mod.monthrange(y, m)[1]
        # Python weekday(): Mon=0..Sun=6 → Sunday-first column index.
        col = ((date(y, m, 1).weekday() + 1) % 7) + 1
        row = cur_row
        for day in range(1, days_in_month + 1):
            c = ws1.cell(row=row, column=col, value=day)
            c.alignment = CENTER
            if date(y, m, day) in occupied:
                c.fill = OCCUPIED_FILL
            col += 1
            if col > 7:
                col = 1
                row += 1
        cur_row = row + (2 if col == 1 else 3)
        m += 1
        if m > 12:
            m, y = 1, y + 1

    # 22-category table: 7-column grid (3 rows of 7 + その他 alone).
    cur_row += 1
    ws1.cell(row=cur_row, column=1,
             value="宿泊者数 国籍別内訳").font = Font(bold=True)
    cur_row += 1
    nat_start = cur_row
    for i, cat in enumerate(NATIONALITY_CATEGORIES):
        block, col = i // 7, (i % 7) + 1
        label_row = nat_start + block * 2
        lc = ws1.cell(row=label_row, column=col, value=cat)
        lc.fill = HEADER_FILL
        lc.alignment = CENTER
        lc.font = Font(bold=True)
        vc = ws1.cell(row=label_row + 1, column=col,
                      value=payload["nationality_counts"][cat])
        vc.alignment = CENTER

    last_nat_row = nat_start + ((len(NATIONALITY_CATEGORIES) - 1) // 7) * 2 + 1
    sum_label_row = last_nat_row + 2
    labels = [("宿泊日数", payload["shukuhaku_nissu"]),
              ("宿泊者数", payload["shukuhakusha_su"]),
              ("延べ人数", payload["nobe_ninzu"])]
    for i, (lbl, val) in enumerate(labels):
        lc = ws1.cell(row=sum_label_row, column=i + 1, value=lbl)
        lc.fill = HEADER_FILL
        lc.alignment = CENTER
        lc.font = Font(bold=True)
        vc = ws1.cell(row=sum_label_row + 1, column=i + 1, value=val)
        vc.alignment = CENTER

    # Loud marker for unresolved nationalities — this workbook must not be
    # mistaken for a finished report while any slot is unresolved.
    if payload["unresolved_count"]:
        c = ws1.cell(row=sum_label_row + 3, column=1,
                     value=T("国籍未確定: %d名 — 提出前に要解決"
                             % payload["unresolved_count"],
                             "UNRESOLVED nationalities: %d — resolve before "
                             "filing" % payload["unresolved_count"]))
        c.font = Font(bold=True, color="9C0006")
        c.fill = ALERT_FILL

    for col_letter in "ABCDEFG":
        ws1.column_dimensions[col_letter].width = 14

    # ── Sheet 2: 明細 (per-reservation detail) ──
    ws2 = wb.create_sheet(title="明細")
    headers = [T("確認コード", "Confirmation code"), T("ゲスト名", "Guest name"),
               T("リスティング", "Listing"), T("ソース", "Source"),
               T("チェックイン", "Check-in"), T("チェックアウト", "Check-out"),
               T("対象開始", "Clipped start"), T("対象終了", "Clipped end"),
               T("泊数(期間内)", "Nights (in period)"),
               T("大人", "Adults"), T("子供", "Children"), T("乳幼児", "Infants"),
               T("宿泊者数", "Guests"), T("延べ人数", "Person-nights"),
               T("ステータス", "Status"), T("国籍", "Nationalities"),
               T("国籍ソース", "Nationality source")]
    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
    for i, r in enumerate(payload["reservations"], 2):
        vals = [r["code"], r["guest_name"], r["listing_name"], r["source"],
                r["checkin"], r["checkout"], r["clip_start"], r["clip_end"],
                r["clipped_nights"], r["adults"], r["children"], r["infants"],
                r["guests"], r["person_nights"], r["status"],
                ", ".join(n if n else T("不明", "UNKNOWN")
                          for n in r["nationalities"]),
                "; ".join(r["nat_sources"])]
        for col, v in enumerate(vals, 1):
            # untrusted text (code/guest/listing/source/nationality strings)
            # is neutralized; numbers and engine-formatted dates pass through.
            ws2.cell(row=i, column=col, value=_safe_cell(v))
        if r["issues"]:  # highlight rows that need review
            for col in range(1, len(headers) + 1):
                ws2.cell(row=i, column=col).fill = OCCUPIED_FILL

    # ── Sheet 3: 確認ログ (review log) ──
    ws3 = wb.create_sheet(title="確認ログ")
    for col, h in enumerate([T("確認コード", "Confirmation code"),
                             T("ゲスト名", "Guest name"),
                             T("宿泊者数", "Guests"),
                             T("問題", "Issues")], 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
    for i, ri in enumerate(payload["review_items"], 2):
        ws3.cell(row=i, column=1, value=_safe_cell(ri["code"]))
        ws3.cell(row=i, column=2, value=_safe_cell(ri["guest_name"]))
        ws3.cell(row=i, column=3, value=ri["guests"])
        ws3.cell(row=i, column=4, value=_safe_cell("; ".join(ri["issues"])))

    # ── Sheet 4 (Kyoto only): 苦情ログ ──
    if payload["kyoto"] and payload["kujo"] is not None:
        ws4 = wb.create_sheet(title="苦情ログ")
        ws4["A1"] = T("苦情の件数", "Number of complaints")
        ws4["A1"].font = Font(bold=True)
        ws4["B1"] = len(payload["kujo"])
        for col, h in enumerate([T("受けた日付", "Date"), T("時刻", "Time"),
                                 T("内容", "Content"),
                                 T("対応状況", "Response status")], 1):
            c = ws4.cell(row=3, column=col, value=h)
            c.font = Font(bold=True)
            c.fill = HEADER_FILL
        for i, e in enumerate(payload["kujo"], 4):
            ws4.cell(row=i, column=1, value=_safe_cell(e["date"]))
            ws4.cell(row=i, column=2, value=_safe_cell(e["time"]))
            ws4.cell(row=i, column=3, value=_safe_cell(e["content"]))
            ws4.cell(row=i, column=4, value=_safe_cell(e["response"]))
        ws4.column_dimensions["C"].width = 50
        ws4.column_dimensions["D"].width = 40

    out_path = os.path.join(output_dir, "%s_report.xlsx" % payload["todokede"])
    wb.save(out_path)
    return out_path


# ───────────────────────── portal CSV ─────────────────────────────────

# CSV COLUMN ORDER (both variants): 届出番号, 報告期間, 宿泊日数, 宿泊者数,
# 延べ人数, the 22 nationality counts (香港 INCLUDED — its omission in the
# 操作手順書's 表2-7 text is a typo; the manual's own Excel screenshot 図2-46
# and the 電子宿泊者名簿 manual's 表1-2 both include it), then 宿泊日.
#
# ⚠ HEADER-ROW CAVEAT: the manuals say row 1 is the 項目名 header but never
# print the canonical header string, and the two official manuals disagree on
# value serialization (操作手順書 v1.7 vs 電子宿泊者名簿 v2.0 — see the
# variant notes below). The headers chosen here follow each manual's own
# vocabulary, but ONLY A LIVE 1-ROW TEST UPLOAD can confirm what the portal
# accepts today. The click guide instructs exactly that test.
SOUSA_HEADER = (["届出番号", "報告期間", "宿泊日数", "宿泊者数", "宿泊延べ人数"]
                + NATIONALITY_CATEGORIES + ["宿泊日"])
MEIBO_HEADER = (["届出番号", "報告対象期間", "宿泊日数", "宿泊者数", "延べ人数"]
                + NATIONALITY_CATEGORIES + ["宿泊日"])


def render_portal_csv(payloads, period, variant, out_path):
    """Write the bulk-upload CSV in cp932 (Windows Shift-JIS) with CRLF.

    variant 'sousa-tejunsho' (default) follows 操作手順書 v1.7:
        届出番号 = 第M…号 ; 報告期間 = 20XX年度XX月～XX月 ;
        宿泊日 = semicolon-joined YYYY-MM-DD dates.
    variant 'denshi-meibo' follows 電子宿泊者名簿 操作説明書 v2.0:
        届出番号 = bare M######### ; 報告対象期間 = YYYY_M-M ;
        宿泊日 = comma+space-joined YYYY/MM/DD dates (the comma forces the
        field to be quoted inside the CSV — whether the portal parser accepts
        quoted fields is UNVERIFIED; prefer the default variant unless a
        test upload proves otherwise).
    One row per 届出番号 × this period. Zero-report properties get 0/0/0
    with an empty 宿泊日 field.
    """
    header = SOUSA_HEADER if variant == "sousa-tejunsho" else MEIBO_HEADER
    # encoding='cp932' raises on unencodable characters (loud, not lossy).
    with open(out_path, "w", encoding="cp932", newline="") as f:
        writer = csv.writer(f, lineterminator="\r\n")
        writer.writerow(header)
        for p in sorted(payloads, key=lambda x: x["todokede"]):
            if variant == "sousa-tejunsho":
                todokede = p["todokede_display"]
                period_label = period["label_portal"]
                dates = ";".join(p["occupied_dates"])
            else:
                todokede = p["todokede"]
                period_label = period["label_meibo"]
                dates = ", ".join(d.replace("-", "/")
                                  for d in p["occupied_dates"])
            row = ([todokede, period_label, p["shukuhaku_nissu"],
                    p["shukuhakusha_su"], p["nobe_ninzu"]]
                   + [p["nationality_counts"][c]
                      for c in NATIONALITY_CATEGORIES]
                   + [dates])
            writer.writerow(row)
    return out_path


# ───────────────────────── click guide ────────────────────────────────


def _compress_days(days):
    """[1,2,3,5,9,10] -> '1-3, 5, 9-10' (compact click list)."""
    if not days:
        return ""
    runs = []
    start = prev = days[0]
    for d in days[1:]:
        if d == prev + 1:
            prev = d
            continue
        runs.append((start, prev))
        start = prev = d
    runs.append((start, prev))
    return ", ".join(str(a) if a == b else "%d-%d" % (a, b) for a, b in runs)


def render_click_guide(payloads, period, out_path):
    """Per-property guide for manual entry on the 事業実績報告登録 screen."""
    L = []
    L.append(T("# 定期報告 入力ガイド — %s" % period["label_portal"],
               "# Teiki-hokoku entry guide — %s" % period["label_portal"]))
    L.append("")
    L.append(T("このガイドは民泊制度運営システムの『事業実績報告登録』画面に手入力する"
               "ための数値一覧です。**提出操作と最終確認は事業者ご本人の責任で行って"
               "ください。**",
               "This guide lists the values to type into the 事業実績報告登録 "
               "screen of the 民泊制度運営システム. **Filing and final "
               "verification are the operator's own responsibility.**"))
    L.append("")
    L.append(T("## 共通の注意", "## Common warnings"))
    L.append(T("- 画面への行き方: ホーム → 届出一覧（受理済）→ 対象届出にチェック → "
               "「事業実績」ボタン（この経路なら届出番号が自動入力されます）。",
               "- Navigation: Home → 届出一覧（受理済） → check the property → "
               "「事業実績」 button (this path auto-fills the 届出番号)."))
    L.append(T("- **報告期間プルダウンの罠**: 既定で選ばれている期間が正しいとは限り"
               "ません。必ず「%s」を選んでください。" % period["label_portal"],
               "- **Period-pulldown trap**: the default selection is often "
               "NOT the right period. Always select \"%s\"."
               % period["label_portal"]))
    L.append(T("- **リセットの罠**: 報告期間を変更すると届出番号以外の全項目が消えます。"
               "期間を最初に選んでから入力してください。",
               "- **Field-reset trap**: changing 報告期間 wipes every field "
               "except 届出番号. Select the period FIRST, then enter data."))
    L.append(T("- 宿泊日数・宿泊者数は画面が自動計算します。下記の値と一致するか確認して"
               "ください。**延べ人数だけは手入力**です。",
               "- 宿泊日数 and 宿泊者数 are auto-calculated on screen; check "
               "they match the values below. **延べ人数 is typed manually.**"))
    L.append(T("- 提出後の修正は報告期限（翌月15日）まで。それ以降は自治体窓口への連絡が"
               "必要です。保存前に必ず見直してください。",
               "- Corrections are possible only until the filing deadline "
               "(the 15th of the following month); after that you must call "
               "your municipality. Review carefully before saving."))
    L.append("")

    for p in sorted(payloads, key=lambda x: x["todokede"]):
        L.append("---")
        L.append("")
        L.append("## %s — %s" % (p["todokede_display"], p["name"]))
        L.append("")
        L.append(T("1. 届出番号: `%s`" % p["todokede_display"],
                   "1. 届出番号: `%s`" % p["todokede_display"]))
        L.append(T("2. 報告期間: 「%s」を選択" % period["label_portal"],
                   "2. 報告期間: select \"%s\"" % period["label_portal"]))
        if p["zero_report"]:
            L.append(T("3. **0件報告**: この期間の宿泊実績はありません。カレンダーは"
                       "何もクリックせず、国籍欄も空のまま、延べ人数に 0 を入力して"
                       "保存してください（宿泊日数・宿泊者数が 0 になっていることを"
                       "確認）。0件でも報告義務があります。",
                       "3. **Zero report**: no stays this period. Click no "
                       "calendar dates, leave the nationality grid empty, "
                       "type 0 for 延べ人数, and save (confirm 宿泊日数 and "
                       "宿泊者数 show 0). A zero report is still mandatory."))
        else:
            L.append(T("3. 宿泊日選択 — 以下の日をクリック:",
                       "3. 宿泊日選択 — click these dates:"))
            by_month = defaultdict(list)
            for ds in p["occupied_dates"]:
                d = date.fromisoformat(ds)
                by_month[(d.year, d.month)].append(d.day)
            for (yy, mm) in sorted(by_month):
                L.append("   - %d年%d月: %s"
                         % (yy, mm, _compress_days(sorted(by_month[(yy, mm)]))))
            L.append(T("4. 宿泊者数国籍別 — 画面の並び順どおりに入力 (0の欄は空のままで可):",
                       "4. Nationality grid — enter in on-screen order "
                       "(boxes with 0 may stay empty):"))
            L.append("")
            L.append("   | %s | %s |" % (T("国籍", "Category"),
                                         T("人数", "Count")))
            L.append("   |---|---|")
            for cat in NATIONALITY_CATEGORIES:
                L.append("   | %s | %d |" % (cat, p["nationality_counts"][cat]))
            L.append("")
            L.append(T("5. 自動計算の確認: 宿泊日数 = **%d** / 宿泊者数 = **%d** "
                       "になるはず。違えばクリック漏れ・入力ミスです。"
                       % (p["shukuhaku_nissu"], p["shukuhakusha_su"]),
                       "5. Cross-check the auto-calculated fields: 宿泊日数 "
                       "should be **%d** and 宿泊者数 **%d**. Any mismatch "
                       "means a missed click or typo."
                       % (p["shukuhaku_nissu"], p["shukuhakusha_su"])))
            L.append(T("6. 延べ人数 (手入力): **%d**" % p["nobe_ninzu"],
                       "6. 延べ人数 (manual entry): **%d**" % p["nobe_ninzu"]))
        if p["unresolved_count"]:
            L.append("")
            L.append(T("> **警告: 国籍未確定 %d名 — この物件は入力しないでください。"
                       "先に国籍を解決してから再生成を。**" % p["unresolved_count"],
                       "> **WARNING: %d unresolved nationalities — do NOT "
                       "file this property yet. Resolve them and re-run.**"
                       % p["unresolved_count"]))
        if p["kyoto"]:
            L.append("")
            L.append(T("### 京都市の追加報告項目 (条例第16条第1項)",
                       "### Kyoto City extra report items (条例16条1項)"))
            if p["kujo"] is None or len(p["kujo"]) == 0:
                L.append(T("- 苦情の件数: **0件** (kujo-log.csv に期間内の記録なし) — "
                           "本当に0件だったかご自身で確認してください。",
                           "- Complaints: **0** (no in-period entries in "
                           "kujo-log.csv) — please verify it was truly zero."))
            else:
                L.append(T("- 苦情の件数: **%d件**" % len(p["kujo"]),
                           "- Complaints: **%d**" % len(p["kujo"])))
                L.append("")
                L.append("   | %s | %s | %s | %s |"
                         % (T("日付", "Date"), T("時刻", "Time"),
                            T("内容", "Content"), T("対応状況", "Response")))
                L.append("   |---|---|---|---|")
                for e in p["kujo"]:
                    L.append("   | %s | %s | %s | %s |"
                             % (e["date"], e["time"], e["content"],
                                e["response"]))
            L.append(T("- **京都市は独自様式（第4号様式）での提出が既定です。**"
                       "国のシステムには苦情欄がないため、少なくとも苦情の状況は"
                       "市の様式・チャネル（メール plb@city.kyoto.lg.jp — 件名の"
                       "先頭に【定期報告】/ FAX 075-251-7235 / 窓口・郵送）で"
                       "提出してください。国のシステム入力だけで足りるかは"
                       "京都市医療衛生センターに確認を — 詳細はスキル同梱の "
                       "references/kyoto.md 参照。",
                       "- **Kyoto City's own form (第4号様式) is the default "
                       "way to file.** The national system has no complaint "
                       "fields, so at minimum the complaint items must go "
                       "through the city's channels (email "
                       "plb@city.kyoto.lg.jp with 【定期報告】 at the head of "
                       "the subject / FAX 075-251-7235 / counter or mail). "
                       "Ask the Kyoto 医療衛生センター whether national-system "
                       "entry alone would suffice — see the bundled "
                       "references/kyoto.md."))
        L.append("")

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(L) + "\n")
    return out_path


# ─────────────────────── validation report ────────────────────────────


def render_validation(payloads, period, stats, findings, running_180,
                      running_sources, config, out_path):
    L = []
    L.append(T("# 検証レポート — %s" % period["key"],
               "# Validation report — %s" % period["key"]))
    L.append("")
    L.append(T("生成日時: %s", "Generated: %s")
             % datetime.now().strftime("%Y-%m-%d %H:%M"))
    if findings.errors:
        status = T("**総合判定: エラーあり — 修正するまで提出しないでください**",
                   "**Overall: ERRORS — do not submit until fixed**")
    elif findings.warnings:
        status = T("**総合判定: 警告あり — 提出前に確認してください**",
                   "**Overall: WARNINGS — review before submitting**")
    else:
        status = T("**総合判定: 問題は検出されませんでした**",
                   "**Overall: no issues detected**")
    L.append("")
    L.append(status)
    L.append("")

    L.append(T("## 入力サマリー", "## Input summary"))
    L.append(T("- 読み込み行数: %d", "- Rows read: %d") % stats["rows"])
    L.append(T("- キャンセル済み (集計除外): %d",
               "- Cancelled (excluded from totals): %d") % stats["cancelled"])
    L.append(T("- ステータス不明 (計上・要確認): %d",
               "- Status unknown (counted, needs review): %d")
             % stats["unknown_status"])
    L.append(T("- 期間外 (対象外): %d", "- Outside the period (skipped): %d")
             % stats["out_of_period"])
    L.append(T("- 不正な行 (スキップ): %d", "- Malformed rows (skipped): %d")
             % stats["bad"])
    L.append(T("- 設定にない届出番号の行 (除外): %d",
               "- Rows with unknown 届出番号 (excluded): %d")
             % stats["unknown_property"])
    if not period["is_standard"]:
        L.append("")
        L.append(T("> 注意: %s は標準の報告期間 (12-1月/2-3月/4-5月/6-7月/8-9月/"
                   "10-11月) ではありません。" % period["key"],
                   "> Note: %s is not a standard reporting period "
                   "(Dec-Jan/Feb-Mar/Apr-May/Jun-Jul/Aug-Sep/Oct-Nov)."
                   % period["key"]))
    L.append("")

    if findings.errors or findings.warnings:
        L.append(T("> 以下の「行N」は、特に断りがない限り %s/normalized_reservations.csv "
                   "の行番号です。行の修正・削除が必要な場合は、ファイルを自分で"
                   "開かなくても、スキルにその旨を伝えれば編集してもらえます。"
                   % period["key"],
                   "> Unless stated otherwise, \"line N\" below refers to "
                   "%s/normalized_reservations.csv. If a row needs fixing or "
                   "removing, you can simply tell the skill — no need to "
                   "open the file yourself." % period["key"]))
        L.append("")

    for title, items in ((T("## エラー", "## Errors"), findings.errors),
                         (T("## 警告", "## Warnings"), findings.warnings),
                         (T("## 情報", "## Info"), findings.infos)):
        L.append(title)
        if items:
            for msg in items:
                L.append("- %s" % msg)
        else:
            L.append(T("- なし", "- none"))
        L.append("")

    L.append(T("## 物件別チェック", "## Per-property checks"))
    for p in sorted(payloads, key=lambda x: x["todokede"]):
        L.append("")
        L.append("### %s — %s" % (p["todokede_display"], p["name"]))
        L.append(T("- 予約数: %d / 宿泊日数: %d / 宿泊者数: %d / 延べ人数: %d",
                   "- Reservations: %d / 宿泊日数: %d / 宿泊者数: %d / "
                   "延べ人数: %d")
                 % (p["reservation_count"], p["shukuhaku_nissu"],
                    p["shukuhakusha_su"], p["nobe_ninzu"]))
        nat_sum = sum(p["nationality_counts"].values())
        ok = nat_sum + p["unresolved_count"] == p["shukuhakusha_su"]
        L.append(T("- 国籍合計チェック: %s (22分類合計 %d + 未確定 %d = 宿泊者数 %d)",
                   "- Nationality sum check: %s (22-category sum %d + "
                   "unresolved %d = guest total %d)")
                 % (T("OK", "OK") if ok else T("不一致 — エラー", "MISMATCH"),
                    nat_sum, p["unresolved_count"], p["shukuhakusha_su"]))
        cal_ok = len(p["occupied_dates"]) == p["shukuhaku_nissu"]
        L.append(T("- カレンダー整合: %s (宿泊日 %d日)",
                   "- Calendar consistency: %s (%d occupied dates)")
                 % (T("OK", "OK") if cal_ok else T("不一致", "MISMATCH"),
                    len(p["occupied_dates"])))

        total = running_180.get(p["todokede"], p["shukuhaku_nissu"])
        L.append(T("- 年度内宿泊日数 (180日ルール): %d日 (4/1正午起算を日付で近似)",
                   "- Fiscal-year total days (180-day cap): %d (noon-Apr-1 "
                   "basis approximated by date)") % total)
        if total > 180:
            L.append(T("  - **エラー級: 180日を超過しています。旅館業法違反の"
                       "おそれ — 直ちに自治体に相談してください**",
                       "  - **CRITICAL: exceeds 180 days — possible 旅館業法 "
                       "violation; contact your municipality immediately**"))
        elif total >= 170:
            L.append(T("  - **警告: 170日以上 — 上限180日まで残りわずかです**",
                       "  - **WARNING: 170+ days — very close to the 180-day "
                       "cap**"))
        elif total >= 120:
            L.append(T("  - 注意: 120日以上 — 年度残り日数を計画してください "
                       "(システムからも注意喚起メールが届く水準)",
                       "  - Caution: 120+ days — plan the rest of the year "
                       "(the portal emails a caution at this level)"))
        if p["kyoto"] and (p["kujo"] is None or len(p["kujo"]) == 0):
            L.append(T("- 京都市: 期間内の苦情記録なし — 0件で報告する前に、本当に"
                       "苦情がなかったか確認してください",
                       "- Kyoto: no in-period complaint entries — before "
                       "reporting 0, verify there really were none"))
        if p["zero_report"]:
            L.append(T("- 0件報告 (宿泊実績なし。0件でも報告義務あり)",
                       "- Zero report (no stays; still legally required)"))
        if p["juri_date"]:
            jd = date.fromisoformat(p["juri_date"])
            if jd > period["start"]:
                L.append(T("- 届出受理日 %s はこの期間の途中/以降です — 受理日以降が"
                           "報告対象" % p["juri_date"],
                           "- 届出受理日 %s falls inside/after this period — "
                           "the duty starts from acceptance" % p["juri_date"]))
    L.append("")

    L.append(T("## 180日集計のデータ範囲", "## 180-day total: data coverage"))
    L.append(T("集計に使ったフォルダ: %s",
               "Folders included in the total: %s")
             % ", ".join(running_sources))
    L.append(T("**この合計は上記フォルダにあるデータだけから計算しています。"
               "年度内に他の宿泊があれば実際の合計はもっと大きくなります。"
               "民泊制度運営システムの届出詳細画面に表示される年間宿泊日数と"
               "必ず照合してください。**",
               "**This total is computed ONLY from the folders listed above. "
               "If the fiscal year had stays not present in these folders, "
               "the true total is higher. Always cross-check against the "
               "annual cumulative days shown on the portal's 届出詳細 "
               "screen.**"))
    L.append("")

    L.append(T("## 提出前の最終確認", "## Final pre-submission notes"))
    L.append(T("- 提出後の修正は報告期限 (翌月15日) まで。それ以降は自治体窓口へ電話。",
               "- Corrections are possible only until the deadline (15th of "
               "the following month); after that, phone your municipality."))
    if config["output"]["portal_csv"]:
        L.append(T("- アップロードCSVの書式 (%s) は公式マニュアル2種で相違があり、"
                   "現行システムでの受理は未確認です。**必ず1行だけの"
                   "テストアップロードで確認**してから全件を投入してください "
                   "(アップロード画面は行ごとのエラーを表示します)。"
                   % config["csv_variant"],
                   "- The upload-CSV serialization (%s) differs between the "
                   "two official manuals and is UNVERIFIED against the "
                   "current system. **Do a 1-row test upload first** (the "
                   "upload screen reports per-row errors) before uploading "
                   "everything." % config["csv_variant"]))
    L.append(T("- このツールはデータを準備するだけです。提出と内容の最終責任は"
               "事業者にあります。",
               "- This tool only PREPARES the data. Filing, and "
               "responsibility for the content, remain with the operator."))
    L.append("")

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(L) + "\n")
    return out_path


# ─────────────────────────────── main ─────────────────────────────────


def main(argv):
    global LANG

    if len(argv) != 3:
        sys.stderr.write(
            "usage: python3 scripts/generate_report.py <workdir> <period>\n"
            "       <period> = YYYY-MM_YYYY-MM (e.g. 2026-06_2026-07)\n")
        sys.exit(2)

    workdir = os.path.abspath(argv[1])
    if not os.path.isdir(workdir):
        fatal("workdir not found: %s" % workdir)
    period = parse_period(argv[2])

    config = load_config(os.path.join(workdir, "minpaku-config.yaml"))
    LANG = config["language"]
    props = config["properties"]

    period_dir = os.path.join(workdir, period["key"])
    output_dir = os.path.join(period_dir, "output")
    os.makedirs(output_dir, exist_ok=True)

    findings = Findings()

    # ── load inputs ──
    records, stats = load_reservations(
        os.path.join(period_dir, "normalized_reservations.csv"),
        period, props, findings)
    overrides = load_overrides(
        os.path.join(workdir, "nationality-overrides.csv"), findings)
    apply_overrides(records, overrides, findings)
    detect_duplicates(records, findings)

    any_kyoto = any(p["kyoto"] for p in props.values())
    kujo_path = os.path.join(period_dir, "kujo-log.csv")
    kujo_entries = []
    if any_kyoto:
        if os.path.isfile(kujo_path):
            kujo_entries = load_kujo_log(kujo_path, period, findings)
            kyoto_count = sum(1 for p in props.values() if p["kyoto"])
            if kyoto_count > 1:
                findings.warning(T(
                    "京都市物件が%d件ありますが kujo-log.csv は1つです — 全京都物件に"
                    "同じ苦情ログが付きます。物件ごとに苦情を分けたい場合は作業フォルダを"
                    "分けてください (v1の制限)" % kyoto_count,
                    "%d Kyoto properties share one kujo-log.csv — the same "
                    "complaint log is attached to every Kyoto property. "
                    "Use separate work folders to separate logs "
                    "(v1 limitation)" % kyoto_count))
        else:
            findings.warning(T(
                "京都市物件がありますが %s/kujo-log.csv が見つかりません — 京都市は"
                "苦情の件数・日時・内容・対応状況の報告が条例で必要です。苦情が0件"
                "だった場合もその確認を推奨します" % period["key"],
                "Kyoto property present but %s/kujo-log.csv not found — "
                "Kyoto requires reporting complaint count/time/content/"
                "response by ordinance. Even for zero complaints, please "
                "confirm" % period["key"]))
    elif os.path.isfile(kujo_path):
        findings.info(T(
            "kujo-log.csv がありますが kyoto:true の物件がないため使用しません",
            "kujo-log.csv found but no property has kyoto:true — ignored"))

    # ── aggregate per property ──
    by_prop = defaultdict(list)
    for rec in records:
        by_prop[rec["todokede"]].append(rec)

    # Informational cross-check: listing names appearing in the CSV that are
    # not declared in the config's listing mapping.
    for bare, recs in by_prop.items():
        declared = {l["listing_name"] for l in props[bare]["listings"]
                    if l["listing_name"]}
        if declared:
            for ln in sorted({r["listing_name"] for r in recs}):
                if ln and ln not in declared:
                    findings.info(T(
                        "%s: リスティング %r は設定に未登録です (集計には含めています)"
                        % (display_todokede(bare), ln),
                        "%s: listing %r not declared in config (still "
                        "counted)" % (display_todokede(bare), ln)))

    payloads = []
    for bare, prop in sorted(props.items()):
        # Properties whose 届出 was accepted after this period have no duty
        # yet — skip, but say so.
        if prop["juri_date"] and prop["juri_date"] > period["end"]:
            findings.info(T(
                "%s: 届出受理日 %s はこの期間より後 — 今回は報告対象外としてスキップ"
                % (prop["todokede_display"], prop["juri_date"]),
                "%s: 届出受理日 %s is after this period — no reporting duty "
                "yet, skipped" % (prop["todokede_display"], prop["juri_date"])))
            continue
        payloads.append(build_property_payload(
            prop, by_prop.get(bare, []), period,
            kujo_entries if prop["kyoto"] else None, findings))

    # ── 180-day fiscal running totals ──
    running_180, running_sources = scan_fiscal_occupancy(
        workdir, period, props, records)
    for p in payloads:
        total = running_180.get(p["todokede"], 0)
        p["fiscal_180"] = {
            "total_days": total,
            "fiscal_year_start": period["fiscal_year_start"].isoformat(),
            "sources": running_sources,
        }
        if total > 180:
            findings.error(T(
                "%s: 年度内宿泊日数 %d日 — 180日上限を超過。自治体に相談してください"
                % (p["todokede_display"], total),
                "%s: fiscal-year total %d days — EXCEEDS the 180-day cap; "
                "contact your municipality" % (p["todokede_display"], total)))
        elif total >= 170:
            findings.warning(T(
                "%s: 年度内宿泊日数 %d日 — 180日上限まで残り%d日"
                % (p["todokede_display"], total, 180 - total),
                "%s: fiscal-year total %d days — only %d left before the "
                "180-day cap" % (p["todokede_display"], total, 180 - total)))
        elif total >= 120:
            findings.info(T(
                "%s: 年度内宿泊日数 %d日 (120日超 — システムの注意喚起メール水準)"
                % (p["todokede_display"], total),
                "%s: fiscal-year total %d days (over 120 — the portal's "
                "caution-mail level)" % (p["todokede_display"], total)))

    # ── write report_data.json (single source of truth) ──
    report_data = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "period": {
            "key": period["key"],
            "start": period["start"].isoformat(),
            "end": period["end"].isoformat(),
            "fiscal_year": period["fiscal_year"],
            "label_portal": period["label_portal"],
            "label_meibo": period["label_meibo"],
            "is_standard": period["is_standard"],
        },
        "operator_name": config["operator_name"],
        "csv_variant": config["csv_variant"],
        "language": config["language"],
        "input_stats": stats,
        "properties": {p["todokede"]: p for p in payloads},
        "validation": {
            "errors": findings.errors,
            "warnings": findings.warnings,
            "info": findings.infos,
        },
    }
    json_path = os.path.join(output_dir, "report_data.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(report_data, f, ensure_ascii=False, indent=2)

    # ── render outputs ──
    written = [json_path]
    if config["output"]["review_xlsx"]:
        for p in payloads:
            written.append(render_property_xlsx(p, period, output_dir))
    if config["output"]["portal_csv"]:
        written.append(render_portal_csv(
            payloads, period, config["csv_variant"],
            os.path.join(output_dir,
                         "teiki-hokoku-upload_%s.csv" % period["key"])))
    if config["output"]["click_guide"]:
        written.append(render_click_guide(
            payloads, period,
            os.path.join(output_dir, "click-guide_%s.md" % period["key"])))
    # The validation report is ALWAYS written — it is the fail-loud channel.
    written.append(render_validation(
        payloads, period, stats, findings, running_180, running_sources,
        config, os.path.join(output_dir,
                             "validation_%s.md" % period["key"])))

    # ── human-readable summary ──
    print("=" * 64)
    print(T("定期報告 生成結果 — %s", "Teiki-hokoku results — %s")
          % period["label_portal"])
    print("=" * 64)
    for p in payloads:
        print("")
        print("%s  %s%s" % (p["todokede_display"], p["name"],
                            T("（0件報告）", " (zero report)")
                            if p["zero_report"] else ""))
        print(T("  予約数: %d / 宿泊日数: %d / 宿泊者数: %d / 延べ人数: %d",
                "  reservations: %d / stay-days: %d / guests: %d / "
                "person-nights: %d")
              % (p["reservation_count"], p["shukuhaku_nissu"],
                 p["shukuhakusha_su"], p["nobe_ninzu"]))
        nz = ", ".join("%s:%d" % (c, n)
                       for c, n in p["nationality_counts"].items() if n)
        if nz:
            print(T("  国籍: %s", "  nationalities: %s") % nz)
        if p["unresolved_count"]:
            print(T("  [要対応] 国籍未確定: %d名",
                    "  [ACTION NEEDED] unresolved nationalities: %d")
                  % p["unresolved_count"])
        print(T("  年度内宿泊日数: %d日 (180日上限)",
                "  fiscal-year days: %d (cap: 180)")
              % p["fiscal_180"]["total_days"])

    print("")
    print(T("出力ファイル:", "Output files:"))
    for path in written:
        print("  " + path)

    print("")
    if findings.errors:
        print(T("[エラー %d件 / 警告 %d件] — validation_%s.md を確認し、修正するまで"
                "提出しないでください。",
                "[%d errors / %d warnings] — see validation_%s.md; do NOT "
                "submit until fixed.")
              % (len(findings.errors), len(findings.warnings), period["key"]))
        sys.exit(1)
    elif findings.warnings:
        print(T("[警告 %d件] — 提出前に validation_%s.md を確認してください。",
                "[%d warnings] — review validation_%s.md before submitting.")
              % (len(findings.warnings), period["key"]))
    else:
        print(T("検証で問題は検出されませんでした。提出前にご自身でも数値を確認して"
                "ください。",
                "No validation issues detected. Still verify the figures "
                "yourself before filing."))
    sys.exit(0)


if __name__ == "__main__":
    main(sys.argv)
