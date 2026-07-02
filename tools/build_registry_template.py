#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_registry_template.py — 宿泊者名簿テンプレート (xlsx) を生成する。

住宅宿泊事業法 施行規則第7条の名簿記載事項（氏名・住所・職業・宿泊日、
日本国内に住所を有しない外国人は国籍・旅券番号）に、運用上あると便利な
列（パスポート写し保管場所・予約コード・備考）を足したテンプレートを
minpaku-teiki-hokoku/assets/guest-registry-template.xlsx に書き出す。

USAGE
    python3 tools/build_registry_template.py [output.xlsx]

引数を省略するとリポジトリ内の既定パス（上記）に保存する。
Google Sheets にそのままインポートできるよう、基本的な書式のみ使う。
依存: openpyxl のみ。
"""

import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# 列定義: (ヘッダー, 記入ルール, 列幅)
# 「宿泊日」は法令上の1項目だが、実務ではチェックイン/チェックアウトの
# 2列に分けたほうが記入も集計も確実なので分割している。
COLUMNS = [
    ("氏名",
     "必須。宿泊者【全員】を1人1行で（代表者だけでは不十分）",
     18),
    ("住所",
     "必須。日本国内に住所がない場合は国名まで書く",
     30),
    ("職業",
     "必須（民泊の名簿では職業欄が必要です）",
     12),
    ("宿泊日（チェックイン）",
     "必須。YYYY-MM-DD 形式（例 2026-04-10）",
     16),
    ("宿泊日（チェックアウト）",
     "必須。YYYY-MM-DD 形式（例 2026-04-13）",
     16),
    ("国籍",
     "日本国内に住所のない外国人は必須。それ以外も分かれば記入"
     "（定期報告の国籍別集計に使えます）",
     14),
    ("旅券番号",
     "日本国内に住所のない外国人は必須（パスポート記載どおり）",
     14),
    ("パスポート写し保管場所",
     "写しの保存場所（フォルダ名・ファイル名など）",
     24),
    ("予約コード",
     "OTAの確認コード（例 Airbnb: HMXXXXXXXX）",
     14),
    ("備考",
     "同行者の続柄・特記事項など",
     20),
]

TITLE = "宿泊者名簿（住宅宿泊事業法 施行規則第7条）"

NOTE = ("記入ルール: 宿泊者全員を1人1行で記入します。名簿は作成日から3年間の"
        "保存義務があります。日本国内に住所を有しない外国人宿泊者は「国籍」"
        "「旅券番号」が必須で、本人確認のうえパスポートの写しを名簿とともに"
        "保管してください（写しを保管する場合、氏名・国籍・旅券番号の記載は"
        "写しで代えられます）。")

HEADER_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3",
                          fill_type="solid")
EXAMPLE_FONT = Font(italic=True, color="808080")

# 行番号（1始まり）: 1=タイトル, 2=注意書き, 3=列ヘッダー, 4=記入ルール,
# 5=記入例（テンプレートのみ）。データはその下から。
HEADER_ROW = 3
RULE_ROW = 4
FIRST_DATA_ROW = 5

EXAMPLE_ROW_VALUES = [
    "（記入例）民泊 花子", "東京都〇〇区〇〇 1-2-3", "会社員",
    "2026-04-10", "2026-04-13", "日本", "", "", "HMEXAMPLE1",
    "この行は消して使ってください",
]


def build_workbook(include_example=True):
    """テンプレートのWorkbookを組み立てて返す。

    include_example=False にすると記入例の行を省く（サンプル名簿の生成など、
    実データを流し込む用途向け）。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "宿泊者名簿"
    n = len(COLUMNS)

    # 1行目: タイトル
    ws.cell(row=1, column=1, value=TITLE).font = Font(size=14, bold=True)

    # 2行目: 注意書き（全列を結合して折り返し表示）
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
    c = ws.cell(row=2, column=1, value=NOTE)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.font = Font(size=9)
    ws.row_dimensions[2].height = 42

    # 3行目: 列ヘッダー / 4行目: 列ごとの記入ルール
    for col, (header, rule, width) in enumerate(COLUMNS, 1):
        hc = ws.cell(row=HEADER_ROW, column=col, value=header)
        hc.font = Font(bold=True)
        hc.fill = HEADER_FILL
        hc.alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
        rc = ws.cell(row=RULE_ROW, column=col, value=rule)
        rc.font = Font(size=8, color="606060")
        rc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[hc.column_letter].width = width
    ws.row_dimensions[RULE_ROW].height = 40

    # 5行目: 記入例（グレー・イタリック）
    if include_example:
        for col, v in enumerate(EXAMPLE_ROW_VALUES, 1):
            c = ws.cell(row=FIRST_DATA_ROW, column=col, value=v)
            c.font = EXAMPLE_FONT

    # ヘッダーと記入ルールを固定表示
    ws.freeze_panes = "A%d" % FIRST_DATA_ROW
    return wb


def main(argv):
    if len(argv) > 2:
        sys.stderr.write("usage: python3 tools/build_registry_template.py "
                         "[output.xlsx]\n")
        return 2
    if len(argv) == 2:
        out_path = argv[1]
    else:
        repo_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        out_path = os.path.join(repo_root, "minpaku-teiki-hokoku", "assets",
                                "guest-registry-template.xlsx")
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    build_workbook(include_example=True).save(out_path)
    print("written: %s" % out_path)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
